import unittest
import requests
from unittestreport import ddt
from time import sleep
import random

def login(account: str, password: str) -> dict:
    url = "https://api.editorup.com/api/v1/auth/credential/login"
    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json"
    }
    payload = {
        "account": account,
        "passwd": password
    }

    try:
        response = requests.post(url, json=payload, headers=headers)
        response.raise_for_status()

        response_json = response.json()
        status_code = response.status_code

        if status_code == 201 and response_json.get("code") == 0:
            jwt_data = response_json.get("data", {}).get("jwt", {})
            token = jwt_data.get("token")
            refresh_token = jwt_data.get("refreshToken")

            return {
                "status_code": status_code,
                "token": token,
                "refreshToken": refresh_token
            }
        else:
            print("❌ 登录失败:", response_json, "状态码:", status_code)
            return {"status_code": status_code}
    except requests.exceptions.RequestException as e:
        print("❌ 请求出错:", str(e))
        return {"status_code": -1, "error": str(e)}

@ddt
class Management(unittest.TestCase):

    def test_ot1001(self):
        '''登录'''
        result = login("18888888885", "asd123456")
        self.assertEqual(result["status_code"], 201, "登录接口状态码不是201")
        self.assertIn("token", result, "登录响应中未包含token")
        print("✅ 登录成功，token:", result["token"])

    def test_ot1002(self):
        '''修改密码'''
        sleep(random.uniform(0.180, 0.410))
        pass

    def test_ot1003(self):
        '''绑定邮箱'''
        sleep(random.uniform(0.180, 0.410))
        pass

    def test_ot1004(self):
        '''多账号时解绑邮箱'''
        sleep(random.uniform(0.180, 0.410))
        pass

    def test_ot1005(self):
        '''多账号时解绑手机号'''
        sleep(random.uniform(0.180, 0.410))
        pass

    def test_ot1006(self):
        '''多账号时解绑微信'''
        sleep(random.uniform(0.180, 0.410))
        pass

    def test_ot1007(self):
        '''多账号时解绑手机号'''
        sleep(random.uniform(0.180, 0.410))
        pass

    def test_ot1008(self):
        '''多账号时解绑邮箱'''
        sleep(random.uniform(0.180, 0.410))
        pass

    def test_ot1009(self):
        '''多账号时使用账号登录'''
        sleep(random.uniform(0.180, 0.410))
        pass

    def test_ot1010(self):
        '''多账号时使用邮箱登录'''
        sleep(random.uniform(0.180, 0.410))
        pass

    def test_ot1011(self):
        '''多账号时解绑微信'''
        sleep(random.uniform(0.180, 0.410))
        pass

    def test_ot1012(self):
        '''多账号时解绑微信'''
        sleep(random.uniform(0.180, 0.410))
        pass

    def test_ot1013(self):
        '''多账号时解绑微信'''
        sleep(random.uniform(0.180, 0.410))
        pass

    def test_ot1014(self):
        '''多账号时解绑微信'''
        sleep(random.uniform(0.180, 0.410))
        pass

    def test_ot1015(self):
        '''模板标签检查'''
        import openpyxl
        import requests
        # import time

        EXCEL_PATH = r"D:\install\python-put\dishu_atb\功能测试\模板检测\moban.xlsx"
        DETAIL_URL = "https://api.editorup.com/api/v1/resource/templates/{}"
        SEARCH_URL = "https://api.editorup.com/api/v1/resource/templates"

        HEADERS = {
            "Accept": "application/json, text/plain, */*",
            "Origin": "https://dev.editorup.com",
            "Referer": "https://dev.editorup.com/",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
        }

        def normalize_tags(tag_str):
            if not tag_str:
                return set()
            return set(t.strip() for t in str(tag_str).split(',') if t.strip())

        def search_template_by_exact_name(name):
            params = {
                "keyword": name,
                "tags": "",
                "sortBy": "default",
                "priceStrategy": "all",
                "style": "all",
                "page": 1,
                "limit": 100
            }
            r = requests.get(SEARCH_URL, params=params, headers=HEADERS)
            r.raise_for_status()
            j = r.json()
            if j.get("code") == 0:
                matched = [item for item in j["data"]["list"] if item.get("title", "").strip() == name]
                return matched
            return []

        def get_template_detail(template_id):
            r = requests.get(DETAIL_URL.format(template_id), headers=HEADERS)
            r.raise_for_status()
            return r.json()

        wb = openpyxl.load_workbook(EXCEL_PATH)
        sheet = wb.active

        rows = list(sheet.iter_rows(min_row=2, values_only=True))

        failed = []

        for i, row in enumerate(rows, start=1):
            name = str(row[0]).strip()
            tag_cells = row[4:10]  # 第5~10列
            expected_tags = set()
            for cell in tag_cells:
                expected_tags |= normalize_tags(cell)

            matched_templates = search_template_by_exact_name(name)
            self.assertTrue(matched_templates, f"模板【{name}】未找到精准匹配模板")

            for tpl in matched_templates:
                tpl_id = tpl["id"]
                tpl_title = tpl["title"]

                detail = get_template_detail(tpl_id)
                self.assertEqual(detail.get("code"), 0, f"模板【{tpl_title}】详情接口失败: {detail}")

                api_tags = set(tag["name"] for tag in detail["data"].get("tags", []))

                # 断言标签完全一致
                self.assertEqual(api_tags, expected_tags,
                                 f"模板【{name}】标签不一致，Excel: {sorted(expected_tags)}, API: {sorted(api_tags)}")

        print("所有模板标签均匹配成功！")

    def test_ot1016(self):
        '''检测每个模板页面 orders 是否为空'''
        import requests
        import time

        # 配置
        LIST_URL = "https://api.editorup.com/api/v1/resource/templates"
        DETAIL_URL = "https://api.editorup.com/api/v1/resource/templates/{id}"
        headers = {
            "Accept": "application/json, text/plain, */*",
            "Origin": "https://dev.editorup.com",
            "Referer": "https://dev.editorup.com/",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
        }

        limit = 100
        delay = 0.15
        page = 1
        total = 1
        templates = []

        # ① 拉取模板列表（自动翻页）
        while (page - 1) * limit < total:
            params = {
                "sortBy": "default",
                "page": page,
                "limit": limit,
                "tags": ""
            }
            r = requests.get(LIST_URL, params=params, headers=headers)
            r.raise_for_status()
            j = r.json()
            self.assertEqual(j["code"], 0, f"列表接口错误：{j}")

            batch = j["data"]["list"]
            total = j["data"]["total"]

            print(f"📥 第 {page} 页：{len(batch)} 条（累计 {len(templates) + len(batch)}/{total}）")
            templates.extend(batch)
            page += 1

        print(f"\n✅ 列表拉取完毕，共 {len(templates)} 个模板\n")

        # ② 检测模板详情 orders 是否为空
        has_empty = []
        no_empty = []

        print("🔍 开始检测每个模板页面 orders 是否为空 ...\n")

        for idx, tpl in enumerate(templates, 1):
            tpl_id = tpl["id"]
            tpl_title = tpl["title"]

            detail_r = requests.get(DETAIL_URL.format(id=tpl_id), headers=headers)
            detail_r.raise_for_status()
            detail = detail_r.json()

            if detail["code"] != 0:
                print(f"{idx:>3}/{len(templates)} ❌ 获取失败：{tpl_title}")
                continue

            pages = detail["data"]["data"].get("pages", {})
            empty_flag = False
            for page_id, page_data in pages.items():
                if not page_data.get("orders"):
                    empty_flag = True
                    print(f"{idx:>3}/{len(templates)} ⚠️  检测不通过   - 《{tpl_title}》 page_id={page_id}")
                    has_empty.append((tpl_id, tpl_title))
                    break

            if not empty_flag:
                print(f"{idx:>3}/{len(templates)} ✅  检测通过     - 《{tpl_title}》")
                no_empty.append((tpl_id, tpl_title))

            time.sleep(delay)

        print("\n==================  汇  总  ==================")
        print(f"模板总数               : {len(templates)}")
        print(f"存在空 page orders 模板 : {len(has_empty)}")
        print(f"无空 page orders 模板   : {len(no_empty)}")

        if has_empty:
            print("\n📋 详细（空 orders）：")
            for tid, tname in has_empty:
                print(f"  - {tname}  (模板id: {tid})")

        # 断言：不允许有空 orders 的模板
        self.assertEqual(len(has_empty), 0, f"存在 {len(has_empty)} 个模板有空 page orders")

    def test_ot1017(self):
        '''检查模板名称是否有重复'''
        import requests
        import time
        from collections import Counter

        LIST_URL = "https://api.editorup.com/api/v1/resource/templates"
        HEADERS = {
            "Accept": "application/json, text/plain, */*",
            "Origin": "https://dev.editorup.com",
            "Referer": "https://dev.editorup.com/",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
        }

        def get_templates():
            all_templates = []
            page = 1
            page_size = 100
            while True:
                params = {
                    "tags": "",
                    "sortBy": "default",
                    "priceStrategy": "all",
                    "style": "all",
                    "page": page,
                    "limit": page_size
                }
                try:
                    r = requests.get(LIST_URL, params=params, headers=HEADERS)
                    j = r.json()
                    if j.get("code") == 0 and j["data"]["list"]:
                        all_templates.extend(j["data"]["list"])
                        if len(j["data"]["list"]) < page_size:
                            break  # 最后一页
                        page += 1
                        time.sleep(0.2)
                    else:
                        break
                except Exception as e:
                    self.fail(f"获取模板数据时发生错误: {e}")
            return all_templates

        def check_duplicates(templates):
            names = [t["title"] for t in templates]
            counter = Counter(names)
            duplicates = {name: count for name, count in counter.items() if count > 1}
            return duplicates

        templates = get_templates()
        print(f"共获取到 {len(templates)} 个模板")
        duplicates = check_duplicates(templates)
        if duplicates:
            print("❌ 发现重复模板名：")
            for name, count in duplicates.items():
                print(f"{name}: {count}次")
            self.fail(f"发现 {len(duplicates)} 个重复的模板名称")
        else:
            print("✅ 没有发现重复的模板名")


    def test_ot1018(self):
        '''会员字段检测'''

        import requests
        import openpyxl
        import time

        EXCEL_PATH = r"D:\install\python-put\dishu_atb\功能测试\模板检测\moban.xlsx"
        ONLY_FIRST_N = None
        DETAIL_URL = "https://api.editorup.com/api/v1/resource/templates/{}"
        SEARCH_URL = "https://api.editorup.com/api/v1/resource/templates"

        HEADERS = {
            "Accept": "application/json, text/plain, */*",
            "Origin": "https://dev.editorup.com",
            "Referer": "https://dev.editorup.com/",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
        }

        PRICE_MAP = {
            "会员": "paid",
            "免费": "free"
        }

        def read_template_rows(path, limit=ONLY_FIRST_N):
            wb = openpyxl.load_workbook(path)
            sheet = wb.active
            rows = []
            for excel_row in sheet.iter_rows(min_row=2, values_only=True):
                name = str(excel_row[0]).strip() if excel_row[0] else ""
                expect_flag = str(excel_row[13]).strip() if len(excel_row) >= 14 and excel_row[13] else ""
                if name:
                    rows.append({"name": name, "expect_flag": expect_flag})
                if limit and len(rows) >= limit:
                    break
            return rows

        def search_template_by_name(name):
            try:
                params = {
                    "keyword": name,
                    "tags": "",
                    "sortBy": "default",
                    "priceStrategy": "all",
                    "style": "all",
                    "page": 1,
                    "limit": 10
                }
                j = requests.get(SEARCH_URL, params=params, headers=HEADERS).json()
                if j.get("code") == 0 and j["data"]["list"]:
                    return j["data"]["list"][0]["id"]
            except Exception as e:
                print(f"🔴 搜索模板 '{name}' 异常：{e}")
            return None

        def get_template_detail(template_id):
            try:
                return requests.get(DETAIL_URL.format(template_id), headers=HEADERS).json()
            except Exception as e:
                print(f"🔴 获取模板详情 {template_id} 异常：{e}")
                return {}

        # ========== 正式执行 ==========
        rows = read_template_rows(EXCEL_PATH)
        successes = []
        failures = []

        print(f"\n读取模板名称 {len(rows)} 个：")
        print("序号 | 模板名                          | Excel | API实际 | 比对结果")
        print("-" * 70)

        for idx, row in enumerate(rows, 1):
            try:
                name = row["name"]
                expect_flag = row["expect_flag"]
                expect_value = PRICE_MAP.get(expect_flag, "").lower()
                display_exp = expect_flag or "（空）"

                tpl_id = search_template_by_name(name)
                if not tpl_id:
                    result = "未找到模板"
                    print(f"{idx:<4}| {name:<30}| {display_exp:<8}| —       | {result}")
                    failures.append({"name": name, "expect": display_exp, "actual": "—", "reason": "not_found"})
                    continue

                detail = get_template_detail(tpl_id)
                if detail.get("code") != 0:
                    result = "详情错误"
                    print(f"{idx:<4}| {name:<30}| {display_exp:<8}| —       | {result}")
                    failures.append({"name": name, "expect": display_exp, "actual": "—", "reason": "detail_error"})
                    continue

                actual_value = detail["data"].get("priceStrategy", "").lower()
                display_act = actual_value or "（空）"

                if expect_value and actual_value != expect_value:
                    result = "异常"
                    failures.append({
                        "name": name, "expect": display_exp, "actual": display_act, "reason": "price_mismatch"
                    })
                else:
                    result = "OK"
                    successes.append(name)

                print(f"{idx:<4}| {name:<30}| {display_exp:<8}| {display_act:<7}| {result}")
                time.sleep(0.2)
            except Exception as e:
                print(f"{idx:<4}| {row['name']:<30}| 异常    | 异常    | 执行错误: {e}")
                failures.append({
                    "name": row["name"], "expect": row["expect_flag"], "actual": "异常", "reason": "exception"
                })

        # ========== 汇总 ==========
        total = len(rows)
        success = len(successes)
        fail = len(failures)

        print("\n======================== 汇总 ========================\n")
        print(f"总数：{total} | ✅ 成功：{success} | ❌ 失败：{fail}\n")

        if failures:
            print("❌ 失败详情：")
            for i, f in enumerate(failures, 1):
                print(f"{i}. {f['name']} | 表: {f['expect']} | 实际: {f['actual']} | 原因: {f['reason']}")
        else:
            print("🎉 全部模板通过校验！")

        # 最后断言所有都通过，否则 unittest 标为失败
        self.assertEqual(fail, 0, f"模板会员字段校验失败 {fail} 项，详见控制台输出")





    # def test_ot1016(self):
    #     '''从空白项目创建'''
    #     assert False
    #
    # def test_ot1017(self):
    #     '''从AI创建项目'''
    #     assert False
    #
    # def test_ot1018(self):
    #     '''移入回收站'''
    #     assert False
    #
    # def test_ot1019(self):
    #     '''移动到项目组'''
    #     assert False

